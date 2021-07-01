#create an excel with data of student database and add values, read excel file , add data into db
import mysql.connector
import xlrd
import openpyxl
import pandas
server = 'XXXXX'
db = 'XXXXXdb'
path = "student.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=5,column=3)
print(cell_obj.value)
for i in range(1,11):
    cell_obj = sheet_obj.cell(row=5,column=i)
    print(cell_obj.value)
mydb = mysql.connector.connect(
    host="local host",
    user="user",
    password="123",
    database="student"
)
cursor = mydb.cursor()
cursor.execute("CREATE DATABASE student")
dbse = mydb.cursor()
dbse.execute("SHOW DATABASE")
for y in dbse:
    print(y)
mydb = mysql.connector.connect(
    host="local host",
    user="user",
    password="123",
    database="student"
)
cursor = mydb.cursor()
cursor.execute("CREATE TABLE student (id INT(10),Name VARCHAR(255),Class INT(25))")
df = pandas.read_excel('student.xlsx')
xl_sheet = xlrd.open_workbook("student.xlsx")
mydb = mysql.connector.connect(
    host="local host",
    user="user",
    password="123",
    database="student"
)
cursor = mydb.cursor()
for s in range(0,1):
    sheet = xl_sheet.sheet_by_index(s)
sql = "INSERT INTO student1 (id,Name,Class) VALUES(%s,%s,%s)"

for r in range(1, sheet.nrows):
   id= sheet.cell(r, 0).value
   Name= sheet.cell(r, 1).value
   Class= sheet.cell(r, 2).value

   values = (id,Name,Class)
   cursor.execute(sql)
   cursor.execute(sql, values)
mycursor = mydb.cursor()
mycursor.execute("SELECT FROM student")
myresult = mycursor.fetchall()
for x in myresult:
    print(x)

mydb.commit()
mydb.close()
